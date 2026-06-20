# -*- coding: utf-8 -*-
import json
from collections import defaultdict,Counter
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
D=json.load(open("_simdata3.json"))
serials=D["serials"];stops=D["stops"];defects=D["defects"];daily=D["daily"];plan=D["plan"];pph=D["pph"]
vibd={d["date"]:d["vib"] for d in daily}
models_by_date=defaultdict(set)
for p in plan: models_by_date[p[0]].add(p[2])

wb=Workbook()
NF=Font(name="Arial");HDR=Font(name="Arial",bold=True,color="FFFFFF")
thin=Side(style="thin",color="D9D9D9");BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
def sh(title,hd,rows,widths,fill="2F5496",numfmt=None):
    ws=wb.create_sheet(title);HF=PatternFill("solid",fgColor=fill)
    for c,h in enumerate(hd,1):
        x=ws.cell(1,c,h);x.font=HDR;x.fill=HF;x.alignment=Alignment(horizontal="center",wrap_text=True);x.border=BORD
    for r,row in enumerate(rows,2):
        for c,v in enumerate(row,1):
            x=ws.cell(r,c,v);x.font=NF;x.border=BORD
            if numfmt and c in numfmt:x.number_format=numfmt[c]
    ws.freeze_panes="A2";ws.auto_filter.ref=f"A1:{get_column_letter(len(hd))}{len(rows)+1}"
    for i,w in enumerate(widths,1):ws.column_dimensions[get_column_letter(i)].width=w
    return ws

# 説明
ws=wb.active;ws.title="説明"
intro=[
 ["製造ライン診断 シミュレーションデータ（整合性保証版 v2）",""],
 ["",""],
 ["概要","架空ライン L-01(大阪工場 JP-OSA)の10稼働日(2026/6/1〜6/12)。1日3〜5機種を順次。1シリアル=1行(計26,812行)。"],
 ["生成","離散イベントシミュレーション。停止・段取り・サイクルから出来高/OEE/累計/仕掛を導出し全シート相互整合(検証済)。"],
 ["切替ダウン","機種切替直後に立上げダウン(SU)を注入。計画外ダウンの約半分が段取り直後に集中(現実準拠)。"],
 ["朝礼/初物","生産開始8:30(朝礼8:00-8:30)。朝一番は初物検査(ライン外)で立上げ遅れ→初回生産は概ね8:45〜8:56。"],
 ["トラブル連動","設備トラブル/停止の間はシリアルが1個も生産されない(停止窓内シリアル0で検証)。復帰時間はばらつきあり。"],
 ["共通キー","拠点-ライン-工程(設備マスタ) / 日付 / 時刻 / 機種 / 材料ロット / 設備ID / シリアルNo で結合可能。"],
 ["",""],
 ["シート","内容"],
 ["計画","日別の機種別計画数・計画開始/終了・標準PPH"],
 ["生産実績_シリアル","1シリアル1行。実績機種/計画機種/機種ズレ/判定/実サイクル/材料ロット"],
 ["停止ログ","停止・段取り・故障の時刻/設備/理由コード/区分/分"],
 ["不良データ","NG明細。発生工程・不良項目・発生/流出区分"],
 ["設備トラブル情報","症状/検知/応急/恒久/再発/影響機種(停止ログと整合)"],
 ["ダウンタイム集計","区分別・設備別のPareto(構成比/累計%)とMTTR(派生)"],
 ["時間別PPH","時間×機種の計画/実績PPH・機種ズレ"],
 ["日次サマリ","OEE(A×P×Q,計算式)/累計(三角図)/投入・完成・仕掛(流動数曲線)/振動"],
 ["設備マスタ","拠点-ライン-工程-設備の3ツリー"],
 ["機種マスタ","標準サイクル/標準PPH"],
 ["ダウンタイム理由コード","停止理由コードの定義(区分・ロス分類)"],
 ["稼働カレンダー","稼働/休・シフト・計画稼働/計画停止"],
 ["標準・手順書一覧","SOP/標準/検査基準/保全基準/材料仕様の台帳(RCA着眼点)"],
 ["4M変化記録","人・機械・材料・方法の変更点"],
 ["作業者割当","日別工程別の担当者と認定レベル(4M:人)"],
 ["センサ_振動","M-12振動トレンドと閾値"],
 ["保全履歴_CMMS","直近6ヶ月の保全WO(bad-actor用,一部欠損)"],
 ["_答え_シナリオ","植え込んだ真因と各スキル想定所見(答え合わせ用)"],
 ["",""],
 ["OEE定義","稼働率A=実働/(実働+停止) 性能P=標準時間/実働 品質Q=良品/実績 OEE=A×P×Q。段取り(CO)・立上げ(SU)・初物検査(FP)は別管理。"],
]
for r,(a,b) in enumerate(intro,1):
    ca=ws.cell(r,1,a);cb=ws.cell(r,2,b)
    ca.font=Font(name="Arial",bold=True,size=(14 if r==1 else 11));cb.font=NF
    cb.alignment=Alignment(wrap_text=True,vertical="top")
ws.column_dimensions["A"].width=20;ws.column_dimensions["B"].width=98

sh("計画",["日付","ライン","機種","計画数","計画開始","計画終了","標準PPH"],plan,[12,8,8,10,10,10,10])
sh("生産実績_シリアル",["シリアルNo","日付","生産時刻","ライン","設備","機種(実績)","計画機種","機種ズレ","判定","実サイクル(秒)","材料ロット","備考"],
   serials,[12,12,20,8,8,11,10,9,7,12,12,12])
sh("停止ログ",["日付","ライン","設備","事象","コード","停止(分)","開始日時","終了時刻","区分"],stops,[12,8,8,20,7,9,18,10,12],fill="C55A11")
sh("不良データ",["シリアルNo","日時","機種","発生工程","発生設備","不良項目","原因区分","流出区分","材料ロット"],defects,[12,18,8,10,10,16,9,12,12],fill="C00000")

# 設備トラブル情報 from stops
agg=defaultdict(lambda:[0,0.0,None])
for s in stops: agg[(s[0],s[2],s[8])][0]+=1;agg[(s[0],s[2],s[8])][1]+=s[5]; \
    agg[(s[0],s[2],s[8])].__setitem__(2, agg[(s[0],s[2],s[8])][2] or s[6])
tr=[];seq=defaultdict(int)
def nid(d):seq[d]+=1;return f"TR-{d.replace('-','')}-{seq[d]:02d}"
for s in sorted(stops,key=lambda x:x[6]):
    if s[4]=="BR":
        v=vibd[s[0]];mdl="/".join(sorted(models_by_date[s[0]]))
        tr.append([nid(s[0]),s[6],s[1],s[2],"突発故障",f"圧入機 異音・振動増({v}mm/s)で停止/荷重ばらつき",round(s[5],1),1,
                   "オペレータ＋振動センサ","治具クッション仮調整で復帰","未(治具予防交換基準なし)","有",mdl,"保全","WO-1305"])
for (d,eq,cat),(cnt,mn,ft) in sorted(agg.items()):
    if cat=="チョコ停" and eq=="M-12" and cnt>=3:
        v=vibd[d];mdl="/".join(sorted(models_by_date[d]))
        tr.append([nid(d),ft,"L-01",eq,"チョコ停多発",f"ワーク詰まり/荷重ばらつき(振動{v}mm/s)",round(mn,1),cnt,"オペレータ","都度復帰","未","有",mdl,"製造",""])
    if cat=="段取り立上げ" and mn>=8:
        mdl="/".join(sorted(models_by_date[d]))
        tr.append([nid(d),ft,"L-01",eq,"段取り立上げ不安定",f"機種切替直後の調整停止(治具摩耗の影響)",round(mn,1),cnt,"段取り作業者","手直しで対応","未(治具更新で解消見込み)","有",mdl,"製造",""])
tr.sort(key=lambda r:r[1])
sh("設備トラブル情報",["トラブルID","発生日時","ライン","設備","区分","現象/症状","停止合計(分)","発生回数","検知方法","応急処置","恒久対策","再発","影響機種","起票","関連WO"],
   tr,[16,18,7,7,14,32,11,9,16,22,24,7,12,8,10],fill="C55A11")

# ダウンタイム集計 (Pareto)
catm=Counter();catc=Counter();eqm=Counter();eqc=Counter()
for s in stops:
    catm[s[8]]+=s[5];catc[s[8]]+=1;eqm[s[2]]+=s[5];eqc[s[2]]+=1
tot=sum(catm.values())
rows=[];cum=0
for k,v in sorted(catm.items(),key=lambda x:-x[1]):
    cum+=v;rows.append([k,catc[k],round(v,1),round(v/tot*100,1),round(cum/tot*100,1)])
ws=sh("ダウンタイム集計",["区分","件数","停止合計(分)","構成比%","累計%"],rows,[16,8,12,10,10],fill="385723")
# second table: 設備別
r0=len(rows)+3
ws.cell(r0,1,"設備別").font=Font(name="Arial",bold=True)
hd2=["設備","件数","停止合計(分)","構成比%","累計%"]
for c,h in enumerate(hd2,1):
    x=ws.cell(r0+1,c,h);x.font=HDR;x.fill=PatternFill("solid",fgColor="385723");x.border=BORD
cum=0;tot2=sum(eqm.values())
for i,(k,v) in enumerate(sorted(eqm.items(),key=lambda x:-x[1])):
    cum+=v
    for c,val in enumerate([k,eqc[k],round(v,1),round(v/tot2*100,1),round(cum/tot2*100,1)],1):
        ws.cell(r0+2+i,c,val).font=NF
# MTTR for 設備故障
brc=sum(1 for s in stops if s[4]=="BR");brm=sum(s[5] for s in stops if s[4]=="BR")
ws.cell(r0+2+len(eqm)+1,1,f"参考: 設備故障 件数 {brc} / 合計 {round(brm)}分 / MTTR {round(brm/max(brc,1),1)}分 (M-12)").font=Font(name="Arial",italic=True)
ws.cell(1,7,"※派生(停止ログ集計)。区分はダウンタイム理由コード参照").font=Font(name="Arial",italic=True,color="808080")

sh("時間別PPH",["日付","時間","計画機種","計画PPH","実績機種","実績PPH","機種ズレ"],pph,[12,8,10,10,12,10,9])

# 日次サマリ (formulas)
ws=wb.create_sheet("日次サマリ")
hd=["日付","計画数","実績数","良品数","不良数","実働(分)","停止(分)","段取り(分)","立上げ(分)","標準時間(分)",
    "稼働率A","性能P","品質Q","OEE","達成率","累計計画","累計実績","投入","累計投入","累計完成","仕掛","振動(mm/s)","初物検査(分)"]
for c,h in enumerate(hd,1):
    x=ws.cell(1,c,h);x.font=HDR;x.fill=PatternFill("solid",fgColor="2F5496");x.alignment=Alignment(horizontal="center",wrap_text=True);x.border=BORD
for i,d in enumerate(daily):
    r=i+2
    for c,v in enumerate([d["date"],d["plan"],d["total"],d["good"],d["defect"],d["operating"],d["stop"],d["co"],d["su"],d["ideal"]],1):
        ws.cell(r,c,v).font=NF
    ws.cell(r,11,f"=F{r}/(F{r}+G{r})");ws.cell(r,12,f"=J{r}/F{r}");ws.cell(r,13,f"=D{r}/C{r}")
    ws.cell(r,14,f"=K{r}*L{r}*M{r}");ws.cell(r,15,f"=C{r}/B{r}")
    ws.cell(r,16,f"=B{r}" if i==0 else f"=P{r-1}+B{r}");ws.cell(r,17,f"=C{r}" if i==0 else f"=Q{r-1}+C{r}")
    ws.cell(r,18,f"=B{r}");ws.cell(r,19,f"=R{r}" if i==0 else f"=S{r-1}+R{r}")
    ws.cell(r,20,f"=D{r}" if i==0 else f"=T{r-1}+D{r}");ws.cell(r,21,f"=S{r}-T{r}");ws.cell(r,22,d["vib"]);ws.cell(r,23,d["fp"])
    for c in range(11,23):ws.cell(r,c).font=NF
    for c in (11,12,13,14,15):ws.cell(r,c).number_format="0.0%"
ws.freeze_panes="B2"
for c in range(1,24):ws.column_dimensions[get_column_letter(c)].width=11
ws.column_dimensions["A"].width=12

# 設備マスタ 3-tree
master=[["JP-OSA","大阪工場","L-01",10,"受入","P-RC","入荷検収","-"],
        ["JP-OSA","大阪工場","L-01",20,"加工","M-08","切削","-"],
        ["JP-OSA","大阪工場","L-01",30,"圧入","M-12","圧入","★制約(ボトルネック)"],
        ["JP-OSA","大阪工場","L-01",40,"組立","M-15","組付","-"],
        ["JP-OSA","大阪工場","L-01",50,"検査","M-20","最終検査","圧入深さ測定なし"],
        ["JP-OSA","大阪工場","L-01",60,"梱包","P-PK","梱包","-"]]
sh("設備マスタ",["拠点コード","拠点名","ライン","工程順","工程","設備ID","役割","制約区分"],master,[10,10,8,8,8,8,14,20],fill="1F4E79")
sh("機種マスタ",["機種","標準サイクル(秒)","標準PPH(個/時)"],[[m,c,round(3600/c)] for m,c in {"A":7.5,"B":10.0,"C":12.0,"D":9.0,"E":11.0}.items()],[10,16,16])
rc=[["BR","設備故障","突発故障(機械停止)","可用性ロス","保全"],
    ["CO","段取り","段取り替え(計画)","段取りロス","製造"],
    ["SU","段取り立上げ","立上げ調整(段取り直後)","段取りロス","製造"],
    ["FP","初物検査","初物検査(ライン外)/朝一番","段取りロス","製造/品質"],
    ["MS","チョコ停","ワーク詰まり/荷重ばらつき","可用性・性能ロス","製造/保全"],
    ["MN","軽微停止","その他軽微停止","可用性ロス","製造"]]
sh("ダウンタイム理由コード",["コード","区分","名称","ロス分類","対応部門"],rc,[8,14,24,16,10],fill="385723")

import datetime as dt
cal=[]
d=dt.date(2026,6,1)
while d<=dt.date(2026,6,12):
    wk="月火水木金土日"[d.weekday()]
    if d.weekday()<5: cal.append([str(d),wk,"稼働","日勤(8:30-)",480,75,"朝礼8:00-8:30 / 休憩45 / 生産開始8:30"])
    else: cal.append([str(d),wk,"休","-",0,0,"休日"])
    d+=dt.timedelta(days=1)
sh("稼働カレンダー",["日付","曜日","区分","シフト","計画稼働(分)","計画停止(分)","備考"],cal,[12,6,8,8,12,12,18],fill="7030A0")

docs=[["SOP-PRESS-01","圧入作業標準","作業手順","圧入/M-12","Rev3","2024-11-10","稼働中","治具点検頻度の記載はあるが交換基準は別文書依存"],
      ["STD-PRESS-COND","圧入条件標準(荷重・速度)","標準","圧入/M-12","Rev2","2023-08-22","稼働中","材料ロット別の条件規定なし(変化点管理の穴)"],
      ["SOP-CO-01","段取り替え手順","作業手順","圧入/M-12","Rev1","2022-05-30","稼働中","SMED(内外段取り分離)未適用"],
      ["QC-INSP-FIN","最終検査基準書","検査基準","検査/M-20","Rev4","2025-01-15","稼働中","外観・主要寸法のみ。圧入深さの測定項目なし(流出経路)"],
      ["PM-STD-01","設備保全標準","保全基準","全般","Rev2","2024-03-04","稼働中","圧入治具クッションの予防交換基準なし(事後保全前提)"],
      ["MAT-SPEC-X","材料仕様書 部品X","仕様","受入/P-RC","Rev5","2026-05-12","稼働中","硬度範囲広め。LOT-2206Bは上限近辺"],
      ["WI-FILL-01","ワーク投入要領","作業手順","圧入/M-12","Rev1","2023-02-18","稼働中","詰まり時の復帰手順を含む"]]
sh("標準・手順書一覧",["文書ID","文書名","種別","対象設備/工程","版","最終改訂","状態","備考(RCA着眼点)"],docs,[14,22,10,14,7,12,8,40],fill="7030A0")
m4=[["2026-06-01","方法","L-01","通常稼働開始(基準)","班長"],
    ["2026-06-03","人","M-12","新規オペレータ配置(OJT中)","班長"],
    ["2026-06-06","材料","圧入部品X","材料ロット変更 LOT-2205A→LOT-2206B(硬度上振れ)","資材"],
    ["2026-06-09","機械","M-12","圧入治具クッション材を暫定調整(応急)","保全"]]
sh("4M変化記録",["日付","区分","対象","変更内容","記録者"],m4,[12,8,12,46,10],fill="548235")
dates=[d["date"] for d in daily];assign=[]
for dd in dates:
    new=dd>="2026-06-03"
    assign.append([dd,"日勤","L-01","圧入/M-12","OP-07","作業者G","訓練中",1,"6/3配置・OP-03がOJT監督"] if new else [dd,"日勤","L-01","圧入/M-12","OP-03","作業者C","熟練",78,""])
    assign.append([dd,"日勤","L-01","加工/M-08","OP-02","作業者B","認定",54,""])
    assign.append([dd,"日勤","L-01","組立/M-15","OP-04","作業者D","認定",36,""])
    assign.append([dd,"日勤","L-01","検査/M-20","OP-05","作業者E","認定",60,""])
    assign.append([dd,"日勤","L-01","班長(全般)","OP-01","作業者A","熟練",120,""])
sh("作業者割当",["日付","シフト","ライン","担当工程/設備","作業者ID","氏名","認定レベル","経験(月)","備考"],assign,[12,8,7,14,10,10,10,9,30],fill="548235")
sh("センサ_振動",["日付","設備","振動(mm/s)","注意閾値","警報閾値","判定"],
   [[d["date"],"M-12",d["vib"],4.5,7.1,("警報" if d["vib"]>=7.1 else "注意" if d["vib"]>=4.5 else "正常")] for d in daily],[12,8,12,10,10,8],fill="BF8F00")
cmms=[["WO-1041","2025-12-08","M-12","圧入治具 増し締め","治具ゆるみ",1.2,18000,"修理"],
 ["WO-1067","2025-12-22","M-15","チェーン給油","",0.5,4000,"PM"],
 ["WO-1093","2026-01-15","M-12","圧入治具クッション交換","治具摩耗",2.5,42000,"修理"],
 ["WO-1110","2026-01-28","M-08","刃物交換","",0.8,12000,"PM"],
 ["WO-1135","2026-02-10","M-12","圧入荷重センサ調整","荷重ばらつき",1.5,22000,"修理"],
 ["WO-1158","2026-02-24","M-20","検査治具校正","",0.6,8000,"PM"],
 ["WO-1177","2026-03-09","M-12","圧入治具 増し締め","治具ゆるみ",1.0,15000,"修理"],
 ["WO-1199","2026-03-23","M-15","軸受交換","軸受異音",3.0,55000,"修理"],
 ["WO-1224","2026-04-06","M-12","圧入治具クッション交換","治具摩耗",2.8,45000,"修理"],
 ["WO-1240","2026-04-19","P-PK","コンベア調整","",0.4,5000,"PM"],
 ["WO-1268","2026-05-04","M-12","圧入ガイド摩耗点検","治具摩耗",1.8,None,"点検"],
 ["WO-1291","2026-05-18","M-08","主軸点検","",1.1,None,"PM"],
 ["WO-1305","2026-05-25","M-12","圧入治具 仮調整","治具ゆるみ",0.9,11000,"修理"]]
sh("保全履歴_CMMS",["WO番号","日付","設備","作業内容","故障コード","ダウン(h)","費用(円)","区分"],cmms,[10,12,8,22,14,10,10,8],fill="833C00")
# 答え
ws=wb.create_sheet("_答え_シナリオ");AF=Font(name="Arial",bold=True,color="7F0000");AFILL=PatternFill("solid",fgColor="FCE4D6")
ans=[["【答え合わせ用】植え込んだ真因と想定所見 — 分析前は閲覧非推奨",""],["",""],
 ["真因(物理)","ボトルネック圧入機 M-12 の圧入治具クッション材摩耗。荷重ばらつき→チョコ停増・圧入深さNGを誘発。"],
 ["引き金(変化点)","6/6 材料ロット変更(LOT-2206B,硬度上振れ)で治具負荷増→摩耗加速(4M変化記録)。"],
 ["人/方法","ロット変更時の条件再設定なし=変化点管理欠落。6/3新人配置は赤ニシン(悪化は6/6以降)。"],
 ["潜在(系統)","治具予防交換基準なし(PM形骸化/事後保全)。最終検査に圧入深さ測定なし(流出)。"],
 ["切替ダウン","機種切替直後の立上げダウン(SU)が劣化日ほど増加。計画外ダウンの約51%が段取り直後に集中。"],
 ["三角図/累計","6/8以降に傾き低下、累計計画と累計実績が乖離。"],
 ["流動数曲線","完成が投入に追いつかず仕掛拡大(6/9以降)。"],
 ["PPH/機種ズレ","劣化日は実績PPH<計画PPH。遅延累積で『前機種を継続』=機種ズレ急増(6/9)。"],
 ["OEE分解","可用性↓(故障/チョコ停)+性能↓(速度)+品質↓(圧入深さNG)の三重。6/9に故障38分。"],
 ["ダウンタイム","区分別Pareto:段取り>立上げ>チョコ停>軽微>設備故障。設備別はM-12突出。"],
 ["bad-actor","M-12が保全費・頻度で突出(CMMS)。治具関連の事後修理反復。"],
 ["RCA到達点","治具摩耗→(発生)圧入深さNG/(流出)検査基準欠落。対策=治具予防交換基準+検査に深さ測定+ロット別条件管理。"],
 ["FMEA/PM","故障モード『圧入治具摩耗→荷重ばらつき』S高O中D低。M-12治具の状態基準PM新設。"]]
for r,(a,b) in enumerate(ans,1):
    ca=ws.cell(r,1,a);cb=ws.cell(r,2,b);ca.font=AF;cb.font=NF;cb.alignment=Alignment(wrap_text=True,vertical="top")
    ca.fill=AFILL; 
    if r==1: cb.fill=AFILL
ws.column_dimensions["A"].width=16;ws.column_dimensions["B"].width=94

order=["説明","計画","生産実績_シリアル","停止ログ","不良データ","設備トラブル情報","ダウンタイム集計","時間別PPH","日次サマリ",
 "設備マスタ","機種マスタ","ダウンタイム理由コード","稼働カレンダー","標準・手順書一覧","4M変化記録","作業者割当","センサ_振動","保全履歴_CMMS","_答え_シナリオ"]
wb._sheets.sort(key=lambda s:order.index(s.title) if s.title in order else 99)
wb.save("manufacturing-line-sim-data.xlsx")
print("saved. sheets",len(wb._sheets))
print([s.title for s in wb._sheets])
